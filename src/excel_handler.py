"""
엑셀 파일 처리 모듈
입력 엑셀 파일 읽기 및 결과 엑셀 파일 쓰기 기능 제공
"""

import openpyxl
from typing import List, Dict, Any
import os
import logging


class ExcelHandler:
    """엑셀 파일 읽기/쓰기 핸들러"""

    def __init__(self, file_path: str):
        """
        Args:
            file_path: 엑셀 파일 경로
        """
        self.file_path = file_path
        self.workbook = None
        self.worksheet = None

    def load(self) -> bool:
        """
        엑셀 파일 로드

        Returns:
            성공 여부
        """
        try:
            if not os.path.exists(self.file_path):
                logging.error(f"파일을 찾을 수 없습니다 - {self.file_path}")
                return False

            self.workbook = openpyxl.load_workbook(self.file_path)
            self.worksheet = self.workbook.active
            logging.info(f"엑셀 파일 로드 완료: {self.file_path}")
            return True
        except Exception as e:
            logging.error(f"엑셀 파일 로드 실패 - {e}")
            return False

    def read_questions(self, success_filter: str = 'all') -> List[Dict[str, Any]]:
        """
        엑셀 파일에서 Question과 Ground Truth 읽기

        Args:
            success_filter: 성공여부 필터 ('all', 'O', 'X')

        Returns:
            질문 데이터 리스트 [{"row": 행번호, "question": 질문, "ground_truth": 정답도메인, "success": 성공여부}, ...]
        """
        if not self.worksheet:
            logging.error("워크시트가 로드되지 않았습니다.")
            return []

        questions = []
        filtered_count = 0

        # 헤더 행 스킵 (1행은 헤더로 가정)
        for row_idx, row in enumerate(self.worksheet.iter_rows(min_row=2, values_only=False), start=2):
            # B열: Question, C열: Ground Truth, E열: 성공여부
            question_cell = row[1]  # B열 (인덱스 1)
            ground_truth_cell = row[2]  # C열 (인덱스 2)
            success_cell = row[4] if len(row) > 4 else None  # E열 (인덱스 4)

            question = question_cell.value if question_cell.value else ""
            ground_truth = ground_truth_cell.value if ground_truth_cell.value else ""
            success = success_cell.value if success_cell and success_cell.value else ""

            # 빈 행은 스킵
            if not question.strip():
                continue

            # 성공여부 필터 적용
            if success_filter != 'all':
                if success.strip() != success_filter:
                    filtered_count += 1
                    continue

            questions.append({
                "row": row_idx,
                "question": question.strip(),
                "ground_truth": ground_truth.strip(),
                "success": success.strip() if success else ""
            })

        if success_filter != 'all':
            logging.info(f"필터 '{success_filter}' 적용: {len(questions)}개 선택, {filtered_count}개 제외")
        logging.info(f"총 {len(questions)}개의 질문을 읽었습니다.")
        return questions

    def write_result(self, row: int, classified_domain: str, success: str, opinion: str, opinion_category: str = ""):
        """
        분류 결과를 엑셀 파일에 쓰기

        Args:
            row: 행 번호
            classified_domain: LLM이 분류한 도메인
            success: 성공 여부 (O/X)
            opinion: 분류 의견
            opinion_category: 분류 의견 구분
        """
        if not self.worksheet:
            logging.error("워크시트가 로드되지 않았습니다.")
            return

        # D열: LLM 도메인 분류 결과
        self.worksheet.cell(row=row, column=4, value=classified_domain)
        # E열: 성공 여부
        self.worksheet.cell(row=row, column=5, value=success)
        # F열: 분류 의견
        self.worksheet.cell(row=row, column=6, value=opinion)
        # G열: 분류 의견 구분
        self.worksheet.cell(row=row, column=7, value=opinion_category)

    def save(self, output_path: str) -> bool:
        """
        결과를 파일로 저장

        Args:
            output_path: 저장할 파일 경로

        Returns:
            성공 여부
        """
        try:
            # 출력 디렉토리가 없으면 생성
            output_dir = os.path.dirname(output_path)
            if output_dir and not os.path.exists(output_dir):
                os.makedirs(output_dir)

            self.workbook.save(output_path)
            logging.info(f"결과 파일 저장 완료: {output_path}")
            return True
        except Exception as e:
            logging.error(f"파일 저장 실패 - {e}")
            return False

    def close(self):
        """워크북 닫기"""
        if self.workbook:
            self.workbook.close()
